#!/usr/bin/env python3

"""
Report on journal article acceptance rates.

Does a lot of in-memory processing, so it helps to run this on a
machine with plenty of RAM.  If this script is invoked with the
`--path` option, it will read the data for the report from the files
in that directory captured on a previous run without that option.

See Jira ticket OCEEBMS-301 for original requirements.
Rewritten for OCEEBMS-569 to use the new EBMS database tables.
"""

import argparse
import datetime
import getpass
import json
import os
import sys
import openpyxl


class States:
    "Dictionary of article state values"

    def __init__(self, path):
        self.values = dict()
        for line in open(f"{path}/states"):
            state_id, state_text_id = json.loads(line.strip())
            self.values[state_text_id] = state_id
        self.ABSTRACT_YES = self.values["passed_bm_review"]
        self.ABSTRACT_NO = self.values["reject_bm_review"]
        self.FULL_TEXT_YES = self.values["passed_full_review"]
        self.FULL_TEXT_NO = self.values["reject_full_review"]
        self.FINAL_DECISION = self.values["final_board_decision"]

    @property
    def wanted(self):
        "These are the states represented by the report"

        return (self.ABSTRACT_YES, self.ABSTRACT_NO,
                self.FULL_TEXT_YES, self.FULL_TEXT_NO,
                self.FINAL_DECISION)


class Counts:
    "Base class for counting occurrences of particular states"

    def __init__(self):
        self.abstract_yes = self.abstract_no = 0
        self.full_text_yes = self.full_text_no = 0
        self.ed_board_yes = self.ed_board_no = 0


class Article(Counts):
    """
    Records the number of times we see a particular state for
    this article in connection with a single board.
    """

    def __init__(self, article_id):
        Counts.__init__(self)
        self.id = article_id

    def map_counts(self, counts):
        """
        No matter how many times we see a particular state for
        this article/board combination, it gets mapped to a
        single count when folding the values into the counts
        for the article's journal.
        """

        counts.num_articles += 1
        if self.abstract_yes:
            counts.abstract_yes += 1
        elif self.abstract_no:
            counts.abstract_no += 1
        if self.full_text_yes:
            counts.full_text_yes += 1
        elif self.full_text_no:
            counts.full_text_no += 1
        if self.ed_board_yes:
            counts.ed_board_yes += 1
        elif self.ed_board_no:
            counts.ed_board_no += 1


class Journal(Counts):
    "Statistics for the articles in this journal considered by one board"

    def __init__(self):
        Counts.__init__(self)
        self.num_articles = 0

    def report(self, sheet, row, title):
        "Write the statistics to a single row in a spreadsheet"

        sheet.cell(row=row, column=1, value=title)
        sheet.cell(row=row, column=2, value=self.num_articles)
        sheet.cell(row=row, column=3, value=self.abstract_yes)
        sheet.cell(row=row, column=4, value=self.abstract_no)
        sheet.cell(row=row, column=5, value=self.full_text_yes)
        sheet.cell(row=row, column=6, value=self.full_text_no)
        sheet.cell(row=row, column=7, value=self.ed_board_yes)
        sheet.cell(row=row, column=8, value=self.ed_board_no)


class Board:
    "One of these for each of the PDQ boards"

    def __init__(self, id, name):
        self.id = id
        self.name = name
        self.not_list = set()
        self.articles = dict()

    def __lt__(self, other):
        return self.name < other.name

    def add_sheet(self, book, control):
        "Create a spreadsheet in one of the two report workbooks"

        if "Complementary" in self.name:
            name = "IACT"
        else:
            name = self.name
        sheet = book.create_sheet(title=name)
        for row in sheet["A1:H1"]:
            for cell in row:
                cell.font = control.bold
                cell.alignment = control.alignment
        sheet.column_dimensions["A"].width = 60
        sheet.cell(row=1, column=1, value="Journal Title")
        sheet.cell(row=1, column=2, value="Total")
        sheet.cell(row=1, column=3, value="Abstract Yes")
        sheet.cell(row=1, column=4, value="Abstract No")
        sheet.cell(row=1, column=5, value="Full-Text Yes")
        sheet.cell(row=1, column=6, value="Full-Text No")
        sheet.cell(row=1, column=7, value="Ed Board Yes")
        sheet.cell(row=1, column=8, value="Ed Board No")
        return sheet

    def report(self, control):
        """Create and populate two spreadsheet for this board's data.

        control - access to the workbooks and style settings

        One of the spreadsheets is for the statistical counts on the
        journals on the list of "don't bother with articles in this
        journal when working on this board's queue" (the "not" list),
        and the other sheet is for all the other journals.
        """

        empty = Journal()
        journals = dict()
        for article_id in self.articles:
            journal_id = control.articles[article_id]
            if journal_id not in journals:
                journals[journal_id] = Journal()
            self.articles[article_id].map_counts(journals[journal_id])
        not_listed = self.add_sheet(control.not_listed, control)
        other = self.add_sheet(control.other, control)
        not_listed_row = other_row = 2
        for journal_id in control.journal_ids:
            title = control.journals[journal_id]
            journal = journals.get(journal_id, empty)
            if journal.num_articles:
                if journal_id in self.not_list:
                    row = not_listed_row
                    not_listed_row += 1
                    sheet = not_listed
                else:
                    row = other_row
                    other_row += 1
                    sheet = other
                journal.report(sheet, row, title)
        sys.stderr.write(f"board {self.name} reported\n")


class Control:
    """Top-level logic for the report."""

    def __init__(self, directory):
        """Gather the values from the directory where they are cached.

        directory - string for the location of the cached data files.
        """

        self.directory = directory
        self.states = States(directory)
        self.boards = dict()
        self.journals = dict()
        self.articles = dict()
        self.decision_values = dict()
        self.board_decisions = dict()
        for line in open(f"{directory}/boards"):
            board_id, name = json.loads(line.strip())
            self.boards[board_id] = Board(board_id, name)
        sys.stderr.write(f"loaded {len(self.boards):d} boards\n")
        count = 0
        for line in open(f"{directory}/not_list"):
            journal_id, board_id = json.loads(line.strip())
            self.boards[board_id].not_list.add(journal_id)
            count += 1
        sys.stderr.write(f"loaded {count:d} not-list directives\n")
        for line in open(f"{directory}/journals"):
            journal_id, journal_title = json.loads(line.strip())
            self.journals[journal_id] = journal_title
        sys.stderr.write(f"loaded {len(self.journals):d} journals\n")
        for line in open(f"{directory}/articles"):
            article_id, journal_id = json.loads(line.strip())
            self.articles[article_id] = journal_id
        sys.stderr.write(f"floaded {len(self.articles):d} articles\n")
        count = 0
        for line in open(f"{directory}/article_boards"):
            article_id, board_id = json.loads(line.strip())
            self.boards[board_id].articles[article_id] = Article(article_id)
            count += 1
        sys.stderr.write(f"loaded {count:d} article/board combos\n")
        for line in open(f"{directory}/decision_values"):
            value_id, value_name = json.loads(line.strip())
            self.decision_values[value_id] = value_name
        msg = f"loaded {len(self.decision_values):d} decision values\n"
        sys.stderr.write(msg)
        count = 0
        for line in open(f"{directory}/board_decisions"):
            article_state_id, decision_value_id = json.loads(line.strip())
            decision_value = self.decision_values.get(decision_value_id)
            if article_state_id not in self.board_decisions:
                self.board_decisions[article_state_id] = set()
            self.board_decisions[article_state_id].add(decision_value)
            count += 1
        sys.stderr.write(f"loaded {count:d} board decisions\n")
        count = 0
        for line in open(f"{directory}/article_states"):
            values = json.loads(line.strip())
            art_state_id, article_id, state_id, board_id = values
            article = self.boards[board_id].articles[article_id]
            if state_id == self.states.ABSTRACT_NO:
                article.abstract_no += 1
            elif state_id == self.states.ABSTRACT_YES:
                article.abstract_yes += 1
            elif state_id == self.states.FULL_TEXT_NO:
                article.full_text_no += 1
            elif state_id == self.states.FULL_TEXT_YES:
                article.full_text_yes += 1
            elif state_id == self.states.FINAL_DECISION:
                board_decisions = self.board_decisions.get(art_state_id)
                if board_decisions:
                    if "Not cited" in board_decisions:
                        article.ed_board_no += 1
                    else:
                        article.ed_board_yes += 1
            count += 1
            sys.stderr.write(f"\rloaded {count:d} article states")
        sys.stderr.write("\n")
        self.journal_ids = list(self.journals.keys())
        self.journal_ids.sort(key=lambda k: self.journals[k])
        sys.stderr.write("data loaded\n")

    def report(self):
        """Generate two workbooks for the report (see Board.report())."""

        opts = dict(horizontal="center", vertical="center", wrap_text=True)
        self.alignment = openpyxl.styles.Alignment(**opts)
        self.bold = openpyxl.styles.Font(bold=True)
        self.not_listed = openpyxl.Workbook()
        self.other = openpyxl.Workbook()
        for sheet in self.not_listed.worksheets:
            self.not_listed.remove(sheet)
        for sheet in self.other.worksheets:
            self.other.remove(sheet)
        for board in sorted(self.boards.values()):
            board.report(self)
        self.not_listed.save(f"{self.directory}/not_listed.xlsx")
        self.other.save(f"{self.directory}/not_not_listed.xlsx")


def fetch(opts):
    """Collect the data from the database and store it to the file system.

    opts - runtime options for the script

    We do it this way so we can tweak the layout of the report by
    reading the stored data into the modified code for the Control
    object, without having to spend time talking to the database all
    over again (which is the lengthier part of the job by quite a bit).

    Return the name of the directory where the values are stored.
    """

    import pymysql
    where = str(datetime.date.today()).replace("-", "")
    try:
        os.mkdir(where)
    except Exception as e:
        print(f"{where}: {e}")
        raise
    start = opts["start"]
    end = opts["end"] + " 23:59:59"
    del opts["start"]
    del opts["end"]
    opts["passwd"] = getpass.getpass(f"password for {opts['user']}: ")
    conn = pymysql.connect(**opts)
    cursor = conn.cursor()
    cursor.execute("SET NAMES utf8")
    cursor.execute(f"USE {opts['db']}")
    cursor.execute("SELECT id, name FROM ebms_board WHERE active = 1")
    with open(f"{where}/boards", "w", encoding="utf-8") as fp:
        rows = cursor.fetchall()
        for row in rows:
            fp.write(f"{json.dumps(tuple(row))}\n")
    sys.stderr.write(f"fetched {len(rows):d} boards\n")
    cursor.execute("""\
SELECT j.source_id, n.not_lists_board
  FROM ebms_journal j
  JOIN ebms_journal__not_lists n
    ON n.entity_id = j.id
 WHERE n.not_lists_start <= NOW()""")
    rows = cursor.fetchall()
    with open(f"{where}/not_list", "w", encoding="utf-8") as fp:
        for row in rows:
            fp.write(f"{json.dumps(tuple(row))}\n")
    sys.stderr.write(f"fetched {len(rows):d} not-list rows\n")
    cursor.execute("SELECT source_id, title from ebms_journal")
    rows = cursor.fetchall()
    with open(f"{where}/journals", "w") as fp:
        for row in rows:
            fp.write(f"{json.dumps(tuple(row))}\n")
    sys.stderr.write(f"fetched {len(rows):d} journals\n")
    cursor.execute("""\
SELECT id, source_journal_id FROM ebms_article
 WHERE import_date BETWEEN %s AND %s""", (start, end))
    articles = set()
    with open(f"{where}/articles", "w", encoding="utf-8") as fp:
        count = 0
        row = cursor.fetchone()
        while row:
            articles.add(row[0])
            fp.write(f"{json.dumps(tuple(row))}\n")
            row = cursor.fetchone()
            count += 1
    sys.stderr.write(f"fetched {count:d} articles\n")
    cursor.execute("""\
SELECT entity_id, field_text_id_value
  FROM taxonomy_term__field_text_id
 WHERE bundle = 'states'""")
    with open(f"{where}/states", "w", encoding="utf-8") as fp:
        rows = cursor.fetchall()
        for row in rows:
            fp.write(f"{json.dumps(tuple(row))}\n")
    sys.stderr.write(f"fetched {len(rows):d} states\n")
    states = States(where)
    cursor.execute("""\
SELECT DISTINCT article, board
           FROM ebms_state
          WHERE active = 1""")
    with open(f"{where}/article_boards", "w", encoding="utf-8") as fp:
        count = 0
        row = cursor.fetchone()
        while row:
            if row[0] in articles:
                fp.write(f"{json.dumps(tuple(row))}\n")
                count += 1
            row = cursor.fetchone()
    sys.stderr.write(f"fetched {count:d} article boards\n")
    wanted = ",".join([str(w) for w in states.wanted])
    cursor.execute(f"""\
SELECT id, article, value, board
  FROM ebms_state
 WHERE active = 1
   AND value IN ({wanted})""")
    article_state_ids = set()
    with open(f"{where}/article_states", "w", encoding="utf-8") as fp:
        row = cursor.fetchone()
        count = 0
        while row:
            if row[1] in articles:
                fp.write(f"{json.dumps(tuple(row))}\n")
                count += 1
                article_state_ids.add(row[0])
            row = cursor.fetchone()
    sys.stderr.write(f"fetched {count:d} article states\n")
    cursor.execute("""\
SELECT tid, name
  FROM taxonomy_term_field_data
 WHERE vid = 'board_decisions'
   AND status = 1""")
    rows = cursor.fetchall()
    with open(f"{where}/decision_values", "w", encoding="utf-8") as fp:
        for row in rows:
            fp.write(f"{json.dumps(tuple(row))}\n")
    sys.stderr.write(f"fetched {len(rows):d} decision values\n")
    cursor.execute("""\
SELECT entity_id, decisions_decision
  FROM ebms_state__decisions""")
    row = cursor.fetchone()
    count = 0
    with open(f"{where}/board_decisions", "w", encoding="utf-8") as fp:
        while row:
            if row[0] in article_state_ids:
                fp.write(f"{json.dumps(tuple(row))}\n")
                count += 1
            row = cursor.fetchone()
    sys.stderr.write(f"fetched {count:d} board decisions\n")
    return where


def main():
    """Collect runtime options and generate the report."""

    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--host")
    group.add_argument("--path")
    parser.add_argument("--port", type=int, default=3661)
    parser.add_argument("--db", default="ebms")
    parser.add_argument("--user", default="ebms")
    parser.add_argument("--start", default="2022-07-01")
    parser.add_argument("--end", default="2023-07-01")
    opts = parser.parse_args()
    if opts.path:
        path = opts.path
    else:
        opts = vars(opts)
        del opts["path"]
        path = fetch(opts)
    control = Control(path)
    control.report()


if __name__ == "__main__":
    """Do nothing if loaded as a module."""
    main()
